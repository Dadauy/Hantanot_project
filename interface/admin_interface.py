import telebot
from messages import default_messages_admin, keyboards_admin
import os
import openpyxl
from database.programma import Programma
from database.inter_party import InterParty
from database.programma_organizators import Programma_Org
import datetime


def proverka(msg):
    if type(msg) is str:
        for c in ["КТЦ Югра-Классик", "КВЦ Югра-Экспо", "Рестораны", "июня"]:
            if c in msg:
                return True
        return False
    return False


def test(bot, message, db_sess):
    file_info = bot.get_file(message.document.file_id)
    downloaded_file = bot.download_file(file_info.file_path)
    src = os.path.abspath(message.document.file_name)
    with open(src, 'wb') as new_file:
        new_file.write(downloaded_file)
    book = openpyxl.open(src)
    sheet = book.active
    # db_sess.query(Programma).delete()
    # db_sess.query(InterParty).delete()
    db_sess.commit()
    i = 2
    while i < sheet.max_row + 1:
        if "июня" in sheet[i][0].value:
            day = int(sheet[i][0].value.split(" ")[0])
            i += 1
            while i < sheet.max_row + 1:
                if "КТЦ Югра-Классик" in sheet[i][0].value:
                    i += 1
                    place_2 = ""
                    while i < sheet.max_row + 1:
                        if proverka(sheet[i][0].value):
                            break
                        flag_date = False
                        for j in "1234567890":
                            if j in sheet[i][0].value:
                                flag_date = True
                        if flag_date is True:
                            moder = sheet[i][2].value
                            n = i + 1
                            while sheet[n][0].value is None and n < sheet.max_row + 1:
                                moder += "#" + sheet[n][2].value
                                n += 1
                            st = (sheet[i][0].value).split(" – ")
                            h = int((st[0].split('.'))[0])
                            m = int((st[0].split('.'))[1])
                            h2 = int((st[1].split('.'))[0])
                            m2 = int((st[1].split('.'))[1])
                            adm = Programma(
                                name=sheet[i][1].value,
                                moder=moder,
                                quest=sheet[i][3].value,
                                date_start=datetime.datetime(2023, 6, day, h, m),
                                date_finish=datetime.datetime(2023, 6, day, h2, m2),
                                place="КТЦ Югра-Классик",
                                place_2=place_2,
                            )
                            db_sess.add(adm)
                            db_sess.commit()
                            i = n
                        if flag_date is False:
                            place_2 = sheet[i][0].value
                            i += 1
                elif "КВЦ Югра-Экспо" in sheet[i][0].value:
                    i += 1
                    place_2 = ""
                    while i < sheet.max_row + 1:
                        if proverka(sheet[i][0].value):
                            break
                        flag_date = False
                        for j in "1234567890":
                            if sheet[i][0].value == None:
                                continue #ЭТО УСЛОВИЕ ДОВАИЛ КОСТЯ!!!
                            if j in sheet[i][0].value:
                                flag_date = True
                        if flag_date:
                            moder = sheet[i][2].value
                            n = i
                            n += 1
                            while sheet[n][0].value is None and n < sheet.max_row + 1:
                                if sheet[n][2].value != None: #ЭТО ДОБАВИЛ КОСТЯ!!!!
                                    moder += "#" + sheet[n][2].value
                                n += 1
                            st = (sheet[i][0].value).split(" – ")
                            h = int((st[0].split('.'))[0])
                            m = int((st[0].split('.'))[1])
                            h2 = int((st[1].split('.'))[0])
                            m2 = int((st[1].split('.'))[1])
                            adm = Programma(
                                name=sheet[i][1].value,
                                moder=moder,
                                quest=sheet[i][3].value,
                                date_start=datetime.datetime(2023, 6, day, h, m),
                                date_finish=datetime.datetime(2023, 6, day, h2, m2),
                                place="КВЦ Югра-Экспо",
                                place_2=place_2,
                            )
                            db_sess.add(adm)
                            db_sess.commit()
                            i = n
                        else:
                            place_2 = sheet[i][0].value
                            i += 1
                elif "Рестораны" in sheet[i][0].value:
                    i += 1
                    place_2 = ""
                    while i < sheet.max_row + 1:
                        if proverka(sheet[i][0].value):
                            break
                        flag_date = False
                        for j in "1234567890":
                            if j in sheet[i][0].value:
                                flag_date = True
                        if flag_date:
                            st = (sheet[i][0].value).split(" – ")
                            h = int((st[0].split('.'))[0])
                            m = int((st[0].split('.'))[1])
                            adm = InterParty(
                                date_start=datetime.datetime(2023, 6, day, h, m),
                                comment=sheet[i][1].value,
                                man_now=0,
                                man_max=int(sheet[i][2].value),
                                place=place_2
                            )
                            db_sess.add(adm)
                            db_sess.commit()
                            i += 1
                        else:
                            place_2 = sheet[i][0].value
                            i += 1
                else:
                    break
        else:
            break


def test2(bot, message, db_sess):
    file_info = bot.get_file(message.document.file_id)
    downloaded_file = bot.download_file(file_info.file_path)
    src = os.path.abspath(message.document.file_name)
    with open(src, 'wb') as new_file:
        new_file.write(downloaded_file)
    book = openpyxl.open(src)
    sheet = book.active
    # db_sess.query(Programma_Org).delete()
    db_sess.commit()
    i = 1
    while i < sheet.max_row + 1:
        if "июня" in sheet[i][0].value:
            day = int(sheet[i][0].value.split(" ")[0])
            i += 1
            while i < sheet.max_row + 1:
                if "КТЦ Югра-Классик" in sheet[i][0].value:
                    i += 1
                    place_2 = ""
                    while i < sheet.max_row + 1:
                        if proverka(sheet[i][0].value):
                            break
                        flag_date = False
                        for j in "1234567890":
                            if j in sheet[i][0].value:
                                flag_date = True
                        if flag_date is True:
                            moder = (sheet[i][2].value if not sheet[i][2].value is None else "")
                            number = str(sheet[i][3].value)
                            n = i + 1
                            while sheet[n][0].value is None and n < sheet.max_row + 1:
                                moder += "#" + (sheet[n][2].value if not sheet[n][2].value is None else "")
                                number += "#" + str(sheet[n][3].value)
                                n += 1
                            st = (sheet[i][0].value).split(" – ")
                            h = int((st[0].split('.'))[0])
                            m = int((st[0].split('.'))[1])
                            h2 = int((st[1].split('.'))[0])
                            m2 = int((st[1].split('.'))[1])
                            adm = Programma_Org(
                                name=sheet[i][1].value,
                                moder=moder,
                                numbers=number,
                                date_start=datetime.datetime(2023, 6, day, h, m),
                                date_finish=datetime.datetime(2023, 6, day, h2, m2),
                                place="КТЦ Югра-Классик",
                                place_2=place_2,
                            )
                            db_sess.add(adm)
                            db_sess.commit()
                            i = n
                        if flag_date is False:
                            place_2 = sheet[i][0].value
                            i += 1
                elif "КВЦ Югра-Экспо" in sheet[i][0].value:
                    i += 1
                    place_2 = ""
                    while i < sheet.max_row + 1:
                        if proverka(sheet[i][0].value):
                            break
                        flag_date = False
                        for j in "1234567890":
                            if j in sheet[i][0].value:
                                flag_date = True
                        if flag_date:
                            moder = (sheet[i][2].value if not sheet[i][2].value is None else "")
                            number = str(sheet[i][3].value)
                            n = i
                            n += 1
                            while sheet[n][0].value is None and n < sheet.max_row + 1:
                                moder += "#" + (sheet[n][2].value if not sheet[n][2].value is None else "")
                                number += "#" + str(sheet[n][3].value)
                                n += 1
                            st = (sheet[i][0].value).split(" – ")
                            h = int((st[0].split('.'))[0])
                            m = int((st[0].split('.'))[1])
                            h2 = int((st[1].split('.'))[0])
                            m2 = int((st[1].split('.'))[1])
                            adm = Programma_Org(
                                name=sheet[i][1].value,
                                moder=moder,
                                numbers=number,
                                date_start=datetime.datetime(2023, 6, day, h, m),
                                date_finish=datetime.datetime(2023, 6, day, h2, m2),
                                place="КВЦ Югра-Экспо",
                                place_2=place_2,
                            )
                            db_sess.add(adm)
                            db_sess.commit()
                            i = n
                        else:
                            place_2 = sheet[i][0].value
                            i += 1
                elif "Рестораны" in sheet[i][0].value:
                    i += 1
                    place_2 = ""
                    while i < sheet.max_row + 1:
                        if proverka(sheet[i][0].value):
                            break
                        flag_date = False
                        for j in "1234567890":
                            if j in sheet[i][0].value:
                                flag_date = True
                        if flag_date:
                            moder = (sheet[i][2].value if not sheet[i][2].value is None else "")
                            number = str(sheet[i][3].value)
                            n = i
                            n += 1
                            while sheet[n][0].value is None and n < sheet.max_row + 1:
                                moder += "#" + (sheet[n][2].value if not sheet[n][2].value is None else "")
                                number += "#" + str(sheet[n][3].value)
                                n += 1
                            st = (sheet[i][0].value).split(" – ")
                            h = int((st[0].split('.'))[0])
                            m = int((st[0].split('.'))[1])
                            h2 = int((st[1].split('.'))[0])
                            m2 = int((st[1].split('.'))[1])
                            adm = Programma_Org(
                                name=sheet[i][1].value,
                                moder=moder,
                                numbers=number,
                                date_start=datetime.datetime(2023, 6, day, h, m),
                                date_finish=datetime.datetime(2023, 6, day, h2, m2),
                                place="Рестораны, корабли, концерты, экскурсии",
                                place_2=place_2,
                            )
                            db_sess.add(adm)
                            db_sess.commit()
                            i = n
                        else:
                            place_2 = sheet[i][0].value
                            i += 1
                else:
                    break
        else:
            break


def admin(bot: telebot.TeleBot, message, db_sess):
    bot.send_message(message.chat.id, default_messages_admin.HELLO_MESSAGE,
                     reply_markup=keyboards_admin.func_data())

    @bot.callback_query_handler(func=lambda call: call.data == "add")
    def excel(call):
        bot.send_message(call.message.chat.id, 'Отправьте excel файл!')

    @bot.message_handler(content_types=['document'])
    def handle_file(message):
        test(bot, message, db_sess)
        bot.send_message(message.chat.id, "Все данные обновлены!!!")

    @bot.callback_query_handler(func=lambda call: call.data == "add2")
    def excel(call):
        bot.send_message(call.message.chat.id, 'Отправьте excel файл!')

    @bot.message_handler(content_types=['document'])
    def handle_file(message):
        test2(bot, message, db_sess)
        bot.send_message(message.chat.id, "Все данные обновлены!!!")
